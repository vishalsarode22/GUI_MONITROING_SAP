"""
Fills the client-provided MetroBrands Excel template ("System Monitoring"
sheet) with results from the GUI T-code collector. Does not touch
screenshots -- Excel gets text values only, per requirement.
"""

import shutil
from datetime import datetime
import openpyxl
import re

from datetime import datetime

from core.models import MetricResult
from utils.logger import get_logger

log = get_logger(__name__, "application")

NO_DATA_TEXT = "See attached screenshot for details."

# Maps template row number -> our internal tcode identifier
TCODE_ROW_MAP = {
    4: "SM21",
    5: "ST22",
    6: "SM13",
    7: "SM12",
    8: "SP01",
    9: "SM37",
    10: "SM37_CANCELLED",
    11: "AL08",
    12: "SM51",
    13: "SM66",
    14: "SCOT",
    15: "ST03N",
    16: "SMLG",
    17: "SOST",
    18: "DB12",
    19: "DB01",
    20: "DB02",
    21: "SMQ1",
    22: "SMQ2",
    23: "SM58",
}


def _build_check_text(metric: MetricResult) -> str:
    """Builds the 'Checks' column text, phrased to match the client's
    reference style per T-code, using real extracted data where available."""
    data = metric.extra_data
    tcode = metric.tcode

    if not data:
        return NO_DATA_TEXT

    if tcode == "ST22":
        count = data.get("dump_count")
        if count is not None:
            today_str = datetime.now().strftime("%d.%m.%Y")
            return f"{count} Abap Dumps - {today_str}"

    if tcode == "SM13":
        summary = data.get("update_summary")
        if summary:
            # e.g. "0 Update records found" -> "00 Update record / 00 Error"
            match = re.match(r"(\d+)\s*Update records?\s*found", summary, re.IGNORECASE)
            if match:
                n = int(match.group(1))
                return f"{n:02d} Update record / 00 Error"
        return NO_DATA_TEXT

    if tcode == "SM12":
        count = data.get("lock_count")
        if count is not None:
            return f"{count} selected Lock Entries found"

    if tcode == "SP01":
        count = data.get("spool_count_visible")
        if count is not None:
            return f"{count} Spool requests displayed (visible page)"

    if tcode == "SM37":
        count = data.get("active_jobs")
        if count is not None:
            return f"{count} active job found" if count != 1 else "1 active job found"

    if tcode == "SM37_CANCELLED":
        count = data.get("cancelled_jobs")
        if count is not None:
            return f"{count} Cancelled Jobs"

    if tcode == "AL08":
        summary = data.get("session_summary")
        if summary:
            match = re.match(r"(\d+)\s*user logons with\s*(\d+)\s*back-end sessions", summary, re.IGNORECASE)
            if match:
                sessions = match.group(2)
                return f"{sessions} ABAP sessions"
        return NO_DATA_TEXT

    if tcode in ("SMQ1", "SMQ2"):
        entries = data.get("entries_displayed")
        queues = data.get("queues_displayed")
        if entries is not None and queues is not None:
            return (
                "               Queue Information\n"
                f"Number of Entries Displayed                  {entries}\n"
                f"Number of Queues Displayed                   {queues}"
            )
    if tcode == "SM51":
        count = data.get("instances_started")
        if count is not None:
            return f"{count} Application server{'s' if count != 1 else ''} are active"

    if tcode == "SM58":
        status = data.get("trfc_status")
        if status:
            return status

    if tcode == "SM66":
        running = data.get("running_processes")
        if running is not None:
            return f"{running} work process{'es' if running != 1 else ''} in use (Running)"

    if tcode == "SOST":
        send_requests = data.get("send_requests")
        waiting = data.get("waiting")
        sent = data.get("sent")
        errors = data.get("errors")
        if send_requests is not None:
            return (
                f"{send_requests} Send Requests {waiting} Waiting , "
                f"{sent} sent, {errors} Error"
            )

    if tcode == "ST03N":
        rt = data.get("dialog_avg_response_time_ms")
        if rt is not None:
            return f"Avg. dialog response time: {rt} ms"

    # Fallback: generic key/value formatting for anything not specially handled
    parts = []
    for key, value in data.items():
        if value is None:
            continue
        label = key.replace("_", " ").title()
        parts.append(f"{label}: {value}")
    return "; ".join(parts) if parts else NO_DATA_TEXT


def fill_metrobrands_template(template_path: str, output_path: str,
                               gui_results: list[MetricResult],
                               report_date: str = None) -> str:
    """
    Copies the template, fills System Monitoring sheet rows based on
    gui_results (matched by MetricResult.tcode), saves to output_path.
    """
    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["System Monitoring"]

    if report_date is None:
        report_date = datetime.now().strftime("%d.%m.%Y")
    ws["C2"] = report_date

    by_tcode = {m.tcode: m for m in gui_results if m.tcode}

    for row, tcode in TCODE_ROW_MAP.items():
        metric = by_tcode.get(tcode)
        if metric is None:
            ws[f"C{row}"] = "Not collected this cycle"
            continue

        if metric.display_value == "failed":
            ws[f"C{row}"] = f"Collection failed: {metric.detail}"
            continue

        ws[f"C{row}"] = _build_check_text(metric)

    wb.save(output_path)
    log.info(f"MetroBrands Excel template filled: {output_path}")
    return output_path