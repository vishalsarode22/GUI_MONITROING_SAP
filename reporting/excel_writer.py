"""
Maintains a historical Excel log of monitoring cycles.
Appends one row per metric per cycle to reports/<date>/SAP_BASIS_Monitoring.xlsx
Creates the file with headers if it doesn't exist yet.
"""

import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook, load_workbook

from core.models import MonitoringResult
from utils.logger import get_logger

log = get_logger(__name__, "monitoring")

COLUMNS = [
    "Timestamp", "System", "Client", "Metric", "Value", "Threshold_Warning",
    "Threshold_Critical", "Status", "TCode", "Detail", "AI_Severity",
    "Root_Cause", "Recommendation", "Screenshot",
]


def _reports_dir_for_today() -> str:
    base = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "reports"
    )
    today = datetime.now().strftime("%Y-%m-%d")
    path = os.path.join(base, today)
    os.makedirs(path, exist_ok=True)
    return path


def _excel_path_for_today() -> str:
    return os.path.join(_reports_dir_for_today(), "SAP_BASIS_Monitoring.xlsx")


def _ensure_workbook(path: str) -> Workbook:
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "History"
    ws.append(COLUMNS)
    wb.save(path)
    return wb


def append_result_to_excel(result: MonitoringResult, path: str | None = None) -> str:
    """
    Appends one row per metric in the MonitoringResult to the Excel history file.
    Returns the path to the file written.
    """
    if path is None:
        path = _excel_path_for_today()

    wb = _ensure_workbook(path)
    ws = wb["History"] if "History" in wb.sheetnames else wb.active

    ai = result.ai_analysis
    ai_severity = ai.severity if ai else ""
    ai_root_cause = ai.likely_root_cause if ai else ""
    ai_recommendation = "; ".join(ai.recommended_actions) if ai and ai.recommended_actions else ""

    for m in result.metrics:
        ws.append([
            result.cycle_timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            result.system,
            result.client,
            m.name,
            m.value if m.value is not None else m.display_value,
            m.threshold_warning,
            m.threshold_critical,
            m.status.value,
            m.tcode or "",
            m.detail,
            ai_severity,
            ai_root_cause,
            ai_recommendation,
            m.screenshot_path or "",
        ])

    wb.save(path)
    log.info(f"Appended {len(result.metrics)} rows to Excel history: {path}")
    return path